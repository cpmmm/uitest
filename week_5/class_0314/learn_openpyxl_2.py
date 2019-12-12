# -*- coding: utf-8 -*-
'''
@time: 2019/11/9 0009 17:26
@author: chen
@contact: 1171954100@qq.com
@file: learn_openpyxl_2.py
@desc:
        ┏┓　　　┏┓+ +
　　　┏┛┻━━━┛┻┓ + +
　　　┃　　　　　　　┃ 　
　　　┃　　　━　　　┃ ++ + + +
　　 ████━████ ┃+
　　　┃　　　　　　　┃ +
　　　┃　　　┻　　　┃
　　　┃　　　　　　　┃ + +
　　　┗━┓　　　┏━┛
　　　　　┃　　　┃　　　　　　　　　　　
　　　　　┃　　　┃ + + + +
　　　　　┃　　　┃　　　　Codes are far away from bugs with the animal protecting　　　
　　　　　┃　　　┃ + 　　　　神兽保佑,代码无bug　　
　　　　　┃　　　┃
　　　　　┃　　　┃　　+　　　　　　　　　
　　　　　┃　 　　┗━━━┓ + +
　　　　　┃ 　　　　　　　┣┓
　　　　　┃ 　　　　　　　┏┛
　　　　　┗┓┓┏━┳┓┏┛ + + + +
　　　　　　┃┫┫　┃┫┫
　　　　　　┗┻┛　┗┻┛+ + + +

'''

#读诗

from openpyxl import load_workbook
#三步骤
#打开工作簿
wb=load_workbook('py15_lbb.xlsx')

#定位表单
sheet=wb['lbb']

#读值
#第一种方式
# print(sheet.cell(2,1).value)
# print(sheet.cell(2,2).value)
# print(sheet.cell(2,3).value)
# print(sheet.cell(2,4).value)

#第二种方式
# for i in range(1,5):
#     print(sheet.cell(2,i).value)

# print(sheet.max_row)   #row 获取最大行
# print(sheet.max_column)#column  获取最大列

#优化一下
# for i in range(1,sheet.max_column+1):   #取头不取尾，所以要加1
#     print(sheet.cell(2,i).value)

#读取所有值
for i in range(1,sheet.max_row+1):          #控制行
    for j in range(1,sheet.max_column+1):   #控制列
        if sheet.cell(i,j).value:           #if 后面的语句是非空非零数据，才会执行后面的语句，否则不会执行
            print(sheet.cell(i,j).value)

