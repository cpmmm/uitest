# -*- coding: utf-8 -*-
'''
@time: 2019/11/8 0008 17:43
@author: chen
@contact: 1171954100@qq.com
@file: learn_openpyxl.py
@desc:
        ┏┓　　　┏┓+ +
　　　┏┛┻━━━┛┻┓ + +
　　　┃　　　　　　　┃ 　
　　　┃　　　━　　　┃ ++ + + +
　　 ████━████ ┃+
　　　┃　　　　　　　┃ +
　　　┃　　　┻　　　┃
　　　┃　　　　　　　┃ + +
　　　┗━┓　　　┏━┛
　　　　　┃　　　┃　　　　　　　　　　　
　　　　　┃　　　┃ + + + +
　　　　　┃　　　┃　　　　Codes are far away from bugs with the animal protecting　　　
　　　　　┃　　　┃ + 　　　　神兽保佑,代码无bug　　
　　　　　┃　　　┃
　　　　　┃　　　┃　　+　　　　　　　　　
　　　　　┃　 　　┗━━━┓ + +
　　　　　┃ 　　　　　　　┣┓
　　　　　┃ 　　　　　　　┏┛
　　　　　┗┓┓┏━┳┓┏┛ + + + +
　　　　　　┃┫┫　┃┫┫
　　　　　　┗┻┛　┗┻┛+ + + +

'''

#操作Excel 读写
#创建一个Excel：手动  自动代码创建
#创建Excel文件的模块 workbook
#创建Excel文件的模块 load_workbook

from openpyxl import workbook

# wb=workbook.Workbook()#创建一个workbook
# wb.create_sheet('lbb')#创建表单
# # wb.save('py15_lbb.xlsx')   #另存为xlsx文件,如果不指定路径，就会存在跟当前文件同级的路径下,相对路径，一定要进行保存
#
# #如果指定路径，比如我想存在class_0312下，就可以复制0312的路径，绝对路径，一定要进行保存
# wb.save('E:\\python15\week_5\class_0312\py15_lbb.xlsx')
#workbook sheet cell

#开始操作Excel  读写
from openpyxl import load_workbook
#读操作
#三步骤
# #第一步：打开Excel 工作簿---workbook
wb=load_workbook('py15_lbb.xlsx')          #打开工作簿
#
# #第二步：定位到表单
#sheet=wb.get_sheet_by_name('lbb')#废弃版本
sheet=wb['lbb']   #推荐使用
#
# #第三步：定位单元格  获取内容   根据行列坐标来定位单元格再获取值,先行后列
# #行列坐标都必须是数字 a-1 b-2 c-3，以此类推
# res=sheet.cell(1,2).value
# print(res)
# print('res的值是{},res的数据类型是{}'.format(res,type(res)))
# A3=sheet.cell(3,1).value
# print('A3的值是{},A3的数据类型是{}'.format(A3,type(A3)))
#
# C3=sheet.cell(3,3).value
# print('C3的值是{},C3的数据类型是{}'.format(C3,type(C3)))
#
# D3=sheet.cell(3,4).value
# print('D3的值是{},D3的数据类型是{}'.format(D3,type(D3)))
#
# #C4=sheet.cell(4,3).value #字符串类型的字典
# C4=eval(sheet.cell(4,3).value)
# print('C4的值是{},C4的数据类型是{}'.format(C4,type(C4)))
#
#
# D4=sheet.cell(4,4).value
# print('D4的值是{},D4的数据类型是{}'.format(D4,type(D4)))

#得出一个结论：数字还是数字，其他数据类型全是字符串类型

#写入值  保存  循环读值
#写入值
#三步骤
#打开工作簿，定位表单，写入值
#eval()  可以把数据转成python原本可以识别的数据类型

#写入值：两种方式   修改和新增
sheet.cell(5,1).value='白日依山尽'     #赋值运算  1
sheet.cell(5,2,'黄河入海流')           #写入值   2

#保存  工作簿
# wb.save('python15_lbb.xlsx')       #另存为操作，因为原来没有python15_lbb这个文件
wb.save('py15_lbb.xlsx')             #如果保存到当前Excel，需要关闭Excel，不然会报错

#操作完毕之后，一定要关闭文件
wb.close()

#循环读值