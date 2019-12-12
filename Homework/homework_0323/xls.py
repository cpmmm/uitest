# -*- coding: utf-8 -*-
'''
@time: 2019/11/16 0016 14:59
@author: chen
@contact: 1171954100@qq.com
@file: xls.py
@desc:
        ┏┓　　　┏┓+ +
　　　┏┛┻━━━┛┻┓ + +
　　　┃　　　　　　　┃ 　
　　　┃　　　━　　　┃ ++ + + +
　　 ████━████ ┃+
　　　┃　　　　　　　┃ +
　　　┃　　　┻　　　┃
　　　┃　　　　　　　┃ + +
　　　┗━┓　　　┏━┛
　　　　　┃　　　┃　　　　　　　　　　　
　　　　　┃　　　┃ + + + +
　　　　　┃　　　┃　　　　Codes are far away from bugs with the animal protecting　　　
　　　　　┃　　　┃ + 　　　　神兽保佑,代码无bug　　
　　　　　┃　　　┃
　　　　　┃　　　┃　　+　　　　　　　　　
　　　　　┃　 　　┗━━━┓ + +
　　　　　┃ 　　　　　　　┣┓
　　　　　┃ 　　　　　　　┏┛
　　　　　┗┓┓┏━┳┓┏┛ + + + +
　　　　　　┃┫┫　┃┫┫
　　　　　　┗┻┛　┗┻┛+ + + +

'''
from openpyxl import load_workbook
# import json
def get_data_from_xls():
    '''获取excel的登录信息，保存到列表里面'''
    wb=load_workbook("data.xlsx")

    # sheet =wb['sheet1']   #第一种取表单方式

    sheet=wb.worksheets[0]  #第二章取表单方式

    params=[]   #外面是一个列表，详情可以看param_data.py文件
    for row in range(2,sheet.max_row+1):     #为什么要+1，因为取头不取尾
        user={                               #大列表的是列表或字典，详情请看param_data.py文件
            'login_url':sheet.cell(row,1).value,
            'method':sheet.cell(row,2).value,
            'param': eval(sheet.cell(row, 3).value),
            'expected':sheet.cell(row,4).value,
            'case_id':sheet.cell(row,7).value
            # 'param':json.loads(sheet.cell(row,3).value)
        }
        params.append(user)   #将user添加到params里

    wb.close()                #操作完一定要关闭
    return params

def write_result(row,result):
    wb = load_workbook("data.xlsx")

    # sheet =wb['sheet1']   #第一种取表单方式

    sheet = wb.worksheets[0]  # 第二章取表单方式

    sheet.cell(row,6,result)

    wb.save("data.xlsx")
    wb.close()



