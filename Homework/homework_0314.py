# -*- coding: utf-8 -*-
'''
@time: 2019/11/9 0009 17:51
@author: chen
@contact: 1171954100@qq.com
@file: homework_0314.py
@desc:
        ┏┓　　　┏┓+ +
　　　┏┛┻━━━┛┻┓ + +
　　　┃　　　　　　　┃ 　
　　　┃　　　━　　　┃ ++ + + +
　　 ████━████ ┃+
　　　┃　　　　　　　┃ +
　　　┃　　　┻　　　┃
　　　┃　　　　　　　┃ + +
　　　┗━┓　　　┏━┛
　　　　　┃　　　┃　　　　　　　　　　　
　　　　　┃　　　┃ + + + +
　　　　　┃　　　┃　　　　Codes are far away from bugs with the animal protecting　　　
　　　　　┃　　　┃ + 　　　　神兽保佑,代码无bug　　
　　　　　┃　　　┃
　　　　　┃　　　┃　　+　　　　　　　　　
　　　　　┃　 　　┗━━━┓ + +
　　　　　┃ 　　　　　　　┣┓
　　　　　┃ 　　　　　　　┏┛
　　　　　┗┓┓┏━┳┓┏┛ + + + +
　　　　　　┃┫┫　┃┫┫
　　　　　　┗┻┛　┗┻┛+ + + +

'''
'''#安排一个作业 #写一个类 类里面有2个方法 1）读数据 2）写数据 
#1）读数据可以读取整个Excel里面所有的数据，每一行数据都放到一个子列表里面，所有子列表数据放到一个大列表里面,要求把读取到的数据返回 
#2）写数据可以在Excel里面指定的单元格里面写入指定的值，不需要返回值 
#温馨提示：记得关闭和保存Excel'''
from openpyxl import workbook
from openpyxl import load_workbook


#为什么要一行存在一起？
# 因为测试用例时一行一行的

# wb=load_workbook('py15_lbb.xlsx')
# sheet=wb['lbb']
# all_data=[]                                   #所有子列表数据放到一个大列表里面
# #读取第一行
# for i in range(1,sheet.max_row+1):#取头不取尾
#     sub_list = []                             #每一行数据都放到一个子列表里面
#     sub_list.append(sheet.cell(i,1).value)
#     sub_list.append(sheet.cell(i,2).value)
#     sub_list.append(sheet.cell(i,3).value)
#     sub_list.append(sheet.cell(i,4).value)
#     print(sub_list)
#     all_data.append(sub_list)
# print(all_data)


#写入数据
    #     wb = load_workbook('py15_lbb.xlsx')
    #     sheet=wb['lbb']
    #     #指定的行列值写入指定的值 new_value
    #     sheet.cell(6,1).value='py'
    #     wb.save('py15_lbb.xlsx')
    #     wb.close()


#改成类
class DoExcel:
    #初始化
    def __init__(self,file_name,sheet_name):
        self.file_name=file_name              #对象属性，对象可以调用  self
        self.sheet_name=sheet_name

    # 读取数据
    def read_data(self):
        '''
        :file_name:目标工作簿名称
        :sheet_name:指定的表单名'''
        wb=load_workbook(self.file_name)
        sheet=wb[self.sheet_name]

        all_data=[]                                   #所有子列表数据放到一个大列表里面

        for i in range(1,sheet.max_row+1):#取头不取尾
             sub_list = []                             #每一行数据都放到一个子列表里面
             sub_list.append(sheet.cell(i,1).value)
             sub_list.append(sheet.cell(i,2).value)
             sub_list.append(sheet.cell(i,3).value)
             sub_list.append(sheet.cell(i,4).value)
             # print(sub_list)
             all_data.append(sub_list)

        wb.close()                #关闭文件
        return all_data           #返回测试数据



    #写入数据
    def write_back(self,row,column,new_value):
        wb = load_workbook(self.file_name)
        sheet=wb[self.sheet_name]
        #指定的行列值写入指定的值 new_value
        sheet.cell(row,column).value=new_value
        wb.save(self.file_name)
        wb.close()



if __name__ == '__main__':
    t=DoExcel('py15_lbb.xlsx','lbb')
    t.write_back(6,3,'p')

    all_data=t.read_data()
    print(all_data)





