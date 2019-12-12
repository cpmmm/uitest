# -*- coding: utf-8 -*-
'''
@time: 2019/11/28 0028 18:21
@author: chen
@contact: 1171954100@qq.com
@file: do_excel.py
@desc:
        ┏┓　　　┏┓+ +
　　　┏┛┻━━━┛┻┓ + +
　　　┃　　　　　　　┃ 　
　　　┃　　　━　　　┃ ++ + + +
　　 ████━████ ┃+
　　　┃　　　　　　　┃ +
　　　┃　　　┻　　　┃
　　　┃　　　　　　　┃ + +
　　　┗━┓　　　┏━┛
　　　　　┃　　　┃　　　　　　　　　　　
　　　　　┃　　　┃ + + + +
　　　　　┃　　　┃　　　　Codes are far away from bugs with the animal protecting　　　
　　　　　┃　　　┃ + 　　　　神兽保佑,代码无bug　　
　　　　　┃　　　┃
　　　　　┃　　　┃　　+　　　　　　　　　
　　　　　┃　 　　┗━━━┓ + +
　　　　　┃ 　　　　　　　┣┓
　　　　　┃ 　　　　　　　┏┛
　　　　　┗┓┓┏━┳┓┏┛ + + + +
　　　　　　┃┫┫　┃┫┫
　　　　　　┗┻┛　┗┻┛+ + + +

'''
import openpyxl
from API_6.common import contants

from API_6.common import http_request

from API_6.common import http_request2


class Case:
    '''
    测试用例类，每个测试用例，实际上就是他的一个实例
    '''
    def __init__(self):
        self.case_id=None
        self.title=None
        self.url=None
        self.data=None
        self.method=None
        self.expected=None
        self.actual=None
        self.result=None


class DoExcel:
    '''
    完成excel的读和写
    '''

    def __init__(self,file_name,sheet_name):
        self.file_name=file_name
        self.sheet_name=sheet_name
        self.workbook=openpyxl.load_workbook(file_name)
        self.sheet=self.workbook[self.sheet_name]

    def get_case(self):
        max_row=self.sheet.max_row         #获取最大行数

        cases=[]  #列表，用来存放所有的测试用例
        for r in range(2,max_row+1):       #第一行是列名，从第二行开始，取头不取尾，最大行数+1
            case=Case()        #实例
            case.case_id=self.sheet.cell(row=r,column=1).value
            case.title = self.sheet.cell(row=r, column=2).value
            case.url = self.sheet.cell(row=r, column=3).value
            case.data = self.sheet.cell(row=r, column=4).value
            case.method = self.sheet.cell(row=r, column=5).value
            case.expected = self.sheet.cell(row=r, column=6).value
            cases.append(case)

        self.workbook.close()

        return cases      #返回cases列表


    def write_result(self,row,actual,result):
        sheet=self.workbook[self.sheet_name]
        sheet.cell(row,7).value=actual
        sheet.cell(row,8).value=result
        self.workbook.save(filename=self.file_name)
        self.workbook.close()

if __name__ == '__main__':
    do_excel=DoExcel(contants.case_file,sheet_name='TestAdd')
    cases=do_excel.get_case()
    http_request = http_request2.HTTPRequest2()       #多加了个2
    for case in cases:
        # print(case.case_id)
        # print(case.title)
        # print(case.url)
        # print(case.data)
        # print(case.method)
        # print(case.expected)
        print(case.__dict__)
        resp=http_request.request(case.method,case.url,case.data)
        print(resp.status_code)
        print(resp.text)         #响应文本
        resp_dict=resp.json()
        print(resp_dict)         #返回字典
        actual=resp.text
        if case.expected == actual:       #判断期望结果是否与实际结果一致
            do_excel.write_result(case.case_id+1,actual,'pass')
        else:
            do_excel.write_result(case.case_id+1,actual,'fail')


