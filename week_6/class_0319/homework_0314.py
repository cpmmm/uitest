# -*- coding: utf-8 -*-
'''
@time: 2019/11/9 0009 17:51
@author: chen
@contact: 1171954100@qq.com
@file: homework_0314.py
@desc:
        ┏┓　　　┏┓+ +
　　　┏┛┻━━━┛┻┓ + +
　　　┃　　　　　　　┃ 　
　　　┃　　　━　　　┃ ++ + + +
　　 ████━████ ┃+
　　　┃　　　　　　　┃ +
　　　┃　　　┻　　　┃
　　　┃　　　　　　　┃ + +
　　　┗━┓　　　┏━┛
　　　　　┃　　　┃　　　　　　　　　　　
　　　　　┃　　　┃ + + + +
　　　　　┃　　　┃　　　　Codes are far away from bugs with the animal protecting　　　
　　　　　┃　　　┃ + 　　　　神兽保佑,代码无bug　　
　　　　　┃　　　┃
　　　　　┃　　　┃　　+　　　　　　　　　
　　　　　┃　 　　┗━━━┓ + +
　　　　　┃ 　　　　　　　┣┓
　　　　　┃ 　　　　　　　┏┛
　　　　　┗┓┓┏━┳┓┏┛ + + + +
　　　　　　┃┫┫　┃┫┫
　　　　　　┗┻┛　┗┻┛+ + + +

'''
'''#安排一个作业 #写一个类 类里面有2个方法 1）读数据 2）写数据 
#1）读数据可以读取整个Excel里面所有的数据，每一行数据都放到一个子列表里面，所有子列表数据放到一个大列表里面,要求把读取到的数据返回 
#2）写数据可以在Excel里面指定的单元格里面写入指定的值，不需要返回值 
#温馨提示：记得关闭和保存Excel'''
from openpyxl import workbook
from openpyxl import load_workbook
from week_6.class_0319.read_config import ReadConfig

#为什么要一行存在一起？
# 因为测试用例时一行一行的

# wb=load_workbook('py15_lbb.xlsx')
# sheet=wb['lbb']
# all_data=[]                                   #所有子列表数据放到一个大列表里面
# #读取第一行
# for i in range(1,sheet.max_row+1):#取头不取尾
#     sub_list = []                             #每一行数据都放到一个子列表里面
#     sub_list.append(sheet.cell(i,1).value)
#     sub_list.append(sheet.cell(i,2).value)
#     sub_list.append(sheet.cell(i,3).value)
#     sub_list.append(sheet.cell(i,4).value)
#     print(sub_list)
#     all_data.append(sub_list)
# print(all_data)


#写入数据
    #     wb = load_workbook('py15_lbb.xlsx')
    #     sheet=wb['lbb']
    #     #指定的行列值写入指定的值 new_value
    #     sheet.cell(6,1).value='py'
    #     wb.save('py15_lbb.xlsx')
    #     wb.close()


#改成类
class DoExcel:
    #初始化
    def __init__(self,file_name,sheet_name):
        self.file_name=file_name              #对象属性，对象可以调用  self
        self.sheet_name=sheet_name

    # 读取数据
    def read_data(self):
        '''
        :file_name:目标工作簿名称
        :sheet_name:指定的表单名'''
        wb=load_workbook(self.file_name)
        sheet=wb[self.sheet_name]

        #从配置文件决定读取，决定获取哪些行的数据，作业里不需要这行
        line = ReadConfig('case.conf').get_stringValue('LINENO', 'LINE')

        all_data=[]                                   #所有子列表数据放到一个大列表里面

        for i in range(1,sheet.max_row+1):#取头不取尾
             sub_list = []                             #每一行数据都放到一个子列表里面
             sub_list.append(sheet.cell(i,1).value)
             sub_list.append(sheet.cell(i,2).value)
             sub_list.append(sheet.cell(i,3).value)
             sub_list.append(sheet.cell(i,4).value)
             # print(sub_list)
             all_data.append(sub_list)

        # 后加的，作业里没有
        final_data=[]     #最后返回的数据

        if line=='all':   #读取所有的数据
            final_data=all_data
        else:             #读取指定列表里的指定行数的数据
            for i in eval(line):  #遍历列表里面的行数的数字，eval(line)，是把字符串的line变成列表类型的line
                final_data.append(all_data[i-1])         #添加数据 line列表里面的数字，跟all_data的数据的索引值是-1的关系

        wb.close()                #关闭文件
        # return all_data           #返回测试数据

        return final_data           #这行作业没有，上一行才是作业的


    #写入数据
    def write_back(self,row,column,new_value):
        wb = load_workbook(self.file_name)
        sheet=wb[self.sheet_name]
        #指定的行列值写入指定的值 new_value
        sheet.cell(row,column).value=new_value
        wb.save(self.file_name)
        wb.close()



if __name__ == '__main__':
    # t=DoExcel('py15_lbb.xlsx','lbb')
    # t.write_back(6,3,'p')
    #
    # all_data=t.read_data()
    # print(all_data)

    print(DoExcel('py15_lbb.xlsx','lbb').read_data())   #这一行后加的，作业里不需要这行



